"""Parse lecturer quiz questions from a small CSV or Excel worksheet."""

from __future__ import annotations

import csv
import io
import re
from pathlib import Path
from typing import Iterable


class QuizSpreadsheetError(ValueError):
    """Raised when a quiz spreadsheet cannot be converted safely."""


HEADER_ALIASES = {
    "type": "type",
    "question_type": "type",
    "question": "question",
    "question_text": "question",
    "prompt": "question",
    "correct_answer": "correct_answer",
    "answer": "correct_answer",
    "correct_option": "correct_answer",
    "marks": "marks",
    "mark": "marks",
    "points": "marks",
    "image": "image_url",
    "image_url": "image_url",
    "options": "options",
}

for option_index, option_letter in enumerate("ABCDEF", start=1):
    HEADER_ALIASES[f"option_{option_letter.lower()}"] = f"option_{option_letter.lower()}"
    HEADER_ALIASES[f"option_{option_index}"] = f"option_{option_letter.lower()}"


def _header(value: object) -> str:
    normalized = re.sub(r"[^a-z0-9]+", "_", str(value or "").strip().lower()).strip("_")
    return HEADER_ALIASES.get(normalized, normalized)


def _text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _rows_from_csv(content: bytes) -> list[dict[str, object]]:
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise QuizSpreadsheetError("CSV files must be saved as UTF-8") from exc
    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        raise QuizSpreadsheetError("The CSV file does not contain a header row")
    headers = [_header(name) for name in reader.fieldnames]
    return [dict(zip(headers, row.values())) for row in reader]


def _rows_from_xlsx(content: bytes) -> list[dict[str, object]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise QuizSpreadsheetError("Excel import is temporarily unavailable") from exc
    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise QuizSpreadsheetError("The Excel workbook could not be opened") from exc
    try:
        sheet = workbook.active
        values = sheet.iter_rows(values_only=True)
        header_row = next(values, None)
        while header_row is not None and not any(_text(value) for value in header_row):
            header_row = next(values, None)
        if header_row is None:
            raise QuizSpreadsheetError("The Excel workbook is empty")
        headers = [_header(value) for value in header_row]
        return [dict(zip(headers, row)) for row in values]
    finally:
        workbook.close()


def _split_options(row: dict[str, object]) -> list[str]:
    options = [_text(row.get(f"option_{letter}")) for letter in "abcdef"]
    options = [option for option in options if option]
    packed = _text(row.get("options"))
    if not options and packed:
        separator = "|" if "|" in packed else ";"
        options = [item.strip() for item in packed.split(separator) if item.strip()]
    return options


def _question_type(value: object) -> str:
    normalized = re.sub(r"[\s_-]+", " ", _text(value).lower()).strip()
    aliases = {
        "mcq": "mcq",
        "multiple choice": "mcq",
        "short": "short",
        "short answer": "short",
        "long": "long",
        "long answer": "long",
        "essay": "long",
    }
    return aliases.get(normalized, "")


def _normalize_rows(rows: Iterable[dict[str, object]]) -> list[dict]:
    questions = []
    errors = []
    for row_number, row in enumerate(rows, start=2):
        if not any(_text(value) for value in row.values()):
            continue
        question_type = _question_type(row.get("type"))
        question = _text(row.get("question"))
        options = _split_options(row)
        correct_answer = _text(row.get("correct_answer"))
        image_url = _text(row.get("image_url"))
        try:
            marks = float(_text(row.get("marks")) or "1")
        except ValueError:
            marks = 0

        row_errors = []
        if not question_type:
            row_errors.append("type must be MCQ, short answer, or long answer")
        if not question:
            row_errors.append("question text is required")
        if marks <= 0:
            row_errors.append("marks must be greater than zero")
        if question_type == "mcq":
            if len(options) < 2:
                row_errors.append("an MCQ needs at least two options")
            answer_key = correct_answer.upper()
            if len(answer_key) == 1 and answer_key in "ABCDEF":
                index = ord(answer_key) - ord("A")
                if index < len(options):
                    correct_answer = options[index]
            elif answer_key.isdigit() and 1 <= int(answer_key) <= len(options):
                correct_answer = options[int(answer_key) - 1]
            if not correct_answer:
                row_errors.append("a correct answer is required")
            elif options and correct_answer not in options:
                row_errors.append("the correct answer must match an option or use its letter")
        if question_type == "short":
            if not correct_answer:
                row_errors.append("a correct answer is required")
            elif len(re.findall(r"\b\w+\b", correct_answer)) > 2:
                row_errors.append("the short-answer key can contain at most two words")
        if image_url and not image_url.lower().startswith(("https://", "http://")):
            row_errors.append("image_url must be a complete http(s) URL; otherwise add the image in the builder")
        if row_errors:
            errors.append(f"Row {row_number}: " + "; ".join(row_errors))
            continue
        questions.append({
            "type": question_type,
            "question": question,
            "options": options if question_type == "mcq" else [],
            "correct_answer": correct_answer if question_type in ("mcq", "short") else "",
            "marks": marks,
            "image_url": image_url,
        })
        if len(questions) > 500:
            raise QuizSpreadsheetError("A quiz can contain at most 500 imported questions")
    if errors:
        suffix = f"; plus {len(errors) - 8} more error(s)" if len(errors) > 8 else ""
        raise QuizSpreadsheetError(" | ".join(errors[:8]) + suffix)
    if not questions:
        raise QuizSpreadsheetError("No quiz questions were found in the spreadsheet")
    return questions


def parse_quiz_spreadsheet(filename: str, content: bytes) -> list[dict]:
    extension = Path(filename or "").suffix.lower()
    if extension == ".csv":
        rows = _rows_from_csv(content)
    elif extension == ".xlsx":
        rows = _rows_from_xlsx(content)
    else:
        raise QuizSpreadsheetError("Choose a .csv or .xlsx quiz file")
    return _normalize_rows(rows)
